import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class EntradasExcelExporter:
    def exportar(self, datos, filename):
        """Exportar entradas a Excel con formato profesional"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Entradas"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        centered = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = ["ID", "Usuario ID", "Fecha", "Texto"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
        
        # Datos
        for row, entrada in enumerate(datos, 2):
            for col, valor in enumerate(entrada, 1):
                cell = sheet.cell(row=row, column=col, value=valor)
                if col in [1, 2]:  # ID y Usuario ID
                    cell.alignment = centered
        
        # Ajustar anchos de columna
        column_widths = [8, 12, 12, 40]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = width
        
        # Ajustar altura de filas para texto
        for row in range(2, len(datos) + 2):
            sheet.row_dimensions[row].height = 60
        
        # Formato de texto
        for row in range(2, len(datos) + 2):
            sheet.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        
        # Congelar paneles
        sheet.freeze_panes = "A2"
        
        # Guardar
        workbook.save(filename)