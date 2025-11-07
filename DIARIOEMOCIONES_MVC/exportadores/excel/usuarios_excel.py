import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class UsuariosExcelExporter:
    def exportar(self, datos, filename):
        """Exportar usuarios a Excel con formato profesional"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Usuarios"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        centered = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = ["ID", "Username", "Email", "Fecha Registro"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
        
        # Datos
        for row, usuario in enumerate(datos, 2):
            for col, valor in enumerate(usuario, 1):
                cell = sheet.cell(row=row, column=col, value=valor)
                if col == 1:  # ID
                    cell.alignment = centered
        
        # Ajustar anchos de columna
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Congelar paneles
        sheet.freeze_panes = "A2"
        
        # Guardar
        workbook.save(filename)