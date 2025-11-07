import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class EmocionesExcelExporter:
    def exportar(self, datos, filename):
        """Exportar emociones a Excel con formato profesional"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Emociones"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        centered = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = ["ID", "Nombre", "Emoji"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
        
        # Datos
        for row, emocion in enumerate(datos, 2):
            for col, valor in enumerate(emocion, 1):
                cell = sheet.cell(row=row, column=col, value=valor)
                if col in [1, 3]:  # ID y Emoji
                    cell.alignment = centered
        
        # Ajustar anchos de columna
        column_widths = [8, 20, 10]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = width
        
        # Congelar paneles
        sheet.freeze_panes = "A2"
        
        # Guardar
        workbook.save(filename)