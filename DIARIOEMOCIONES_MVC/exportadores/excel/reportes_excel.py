import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class ReportesExcelExporter:
    def exportar(self, datos, filename):
        """Exportar reportes a Excel con formato profesional"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Reporte Emocional"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        centered = Alignment(horizontal="center", vertical="center")
        
        # Título
        sheet.merge_cells('A1:C1')
        title_cell = sheet.cell(row=1, column=1, value="REPORTE EMOCIONAL - DIARIO DE EMOCIONES")
        title_cell.font = Font(bold=True, size=14, color="7030A0")
        title_cell.alignment = centered
        
        # Encabezados
        headers = ["Emoción", "Frecuencia", "Última Registro"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
        
        # Datos
        for row, reporte in enumerate(datos, 3):
            for col, valor in enumerate(reporte, 1):
                cell = sheet.cell(row=row, column=col, value=valor)
                if col == 2:  # Frecuencia
                    cell.alignment = centered
        
        # Ajustar anchos de columna
        column_widths = [20, 15, 18]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = width
        
        # Estadísticas
        if datos:
            stats_row = len(datos) + 4
            sheet.cell(row=stats_row, column=1, value="TOTAL EMOCIONES:").font = Font(bold=True)
            sheet.cell(row=stats_row, column=2, value=len(datos)).font = Font(bold=True)
        
        # Congelar paneles
        sheet.freeze_panes = "A3"
        
        # Guardar
        workbook.save(filename)